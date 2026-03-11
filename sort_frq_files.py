#!/usr/bin/env python3
"""
sort_frq_files.py — Sort AP FRQ spreadsheets by unit, year, and question.

For each subject folder under this directory, finds the .xlsx file and sorts
its rows so that:
  1. Unit 1 rows come first, Unit 2 next, etc. (extracted from "Unit X: ..." in
     the Unit Topic column)
  2. Within each unit, years run newest → oldest (e.g. 2025, 2024, ..., 2014)
  3. Within each year, questions run lowest → highest (e.g. 1, 2, 3, ...)

Year values like "2006B" (Form B exams) are sorted after "2006" within the
same year. Calculus files use separate "Question Number" and "Letter" columns;
all other subjects use a single "Question" column.

Hyperlinks in the Source column are fully preserved (openpyxl is used directly
instead of pandas, which strips hyperlinks on read).

Usage:
    python3 sort_frq_files.py
"""

import copy
import os
import re
import shutil
import sys
import tempfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent


def unit_sort_key(val) -> float:
    m = re.search(r"Unit\s+(\d+)", str(val))
    return int(m.group(1)) if m else float("inf")


def year_sort_key(val) -> tuple:
    m = re.match(r"^(\d+)(.*)", str(val).strip())
    if m:
        return (int(m.group(1)), m.group(2))
    return (0, "")


def question_sort_key(val) -> tuple:
    s = str(val).strip()
    m = re.match(r"^(\d+)(.*)", s)
    if m:
        return (int(m.group(1)), m.group(2).lower())
    return (float("inf"), s.lower())


def cell_value(cell) -> str:
    """Return cell value as string (empty string for None)."""
    return "" if cell.value is None else str(cell.value)


def sort_file(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[0]

    # Read header row
    header_row = [cell_value(c) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(header_row)}

    is_calculus = "Question Number" in col

    # Snapshot every data row as a list of deep-copied cells (preserves hyperlinks)
    data_rows = []
    for row in ws.iter_rows(min_row=2):
        data_rows.append([copy.copy(c) for c in row])

    # Build sort keys for each row
    def row_keys(cells):
        unit = unit_sort_key(cell_value(cells[col["Unit Topic"]]))
        yr_num, yr_sfx = year_sort_key(cell_value(cells[col["Year"]]))
        if is_calculus:
            q_num = question_sort_key(cell_value(cells[col["Question Number"]]))[0]
            q_sfx = cell_value(cells[col["Letter"]]).strip().lower()
        else:
            q_num, q_sfx = question_sort_key(cell_value(cells[col["Question"]]))
        return (unit, -yr_num, yr_sfx, q_num, q_sfx)

    # Sort: unit asc, year desc (negated), year suffix desc, q_num asc, q_sfx asc
    # Note: year suffix descending means "B" > "" so we negate: use reverse on suffix
    # Actually we want yr_sfx descending ("B" before ""), so we negate the string comparison
    # by wrapping — simplest: sort with a custom key that negates yr_sfx via inversion trick.
    # We'll use a compound key tuple and sort ascending, but flip yr_sfx sign manually:
    def sort_key(cells):
        unit = unit_sort_key(cell_value(cells[col["Unit Topic"]]))
        yr_num, yr_sfx = year_sort_key(cell_value(cells[col["Year"]]))
        if is_calculus:
            q_num = question_sort_key(cell_value(cells[col["Question Number"]]))[0]
            q_sfx = cell_value(cells[col["Letter"]]).strip().lower()
        else:
            q_num, q_sfx = question_sort_key(cell_value(cells[col["Question"]]))
        # year descending: negate yr_num; yr_sfx descending: negate each char ord
        neg_yr_sfx = tuple(-ord(ch) for ch in yr_sfx)
        return (unit, -yr_num, neg_yr_sfx, q_num, q_sfx)

    data_rows.sort(key=sort_key)

    # Write sorted rows back into the worksheet
    for r_idx, cells in enumerate(data_rows, start=2):
        for c_idx, src_cell in enumerate(cells, start=1):
            dst_cell = ws.cell(row=r_idx, column=c_idx)
            dst_cell.value = src_cell.value
            if src_cell.hyperlink:
                dst_cell.hyperlink = copy.copy(src_cell.hyperlink)
            else:
                dst_cell.hyperlink = None
            if src_cell.has_style:
                dst_cell.font = copy.copy(src_cell.font)
                dst_cell.border = copy.copy(src_cell.border)
                dst_cell.fill = copy.copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = copy.copy(src_cell.protection)
                dst_cell.alignment = copy.copy(src_cell.alignment)

    # Atomic write
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=path.suffix, dir=path.parent)
    try:
        os.close(tmp_fd)
        wb.save(tmp_path)
        shutil.move(tmp_path, path)
    except Exception:
        Path(tmp_path).unlink(missing_ok=True)
        raise


def main() -> int:
    files = []
    for subdir in sorted(ROOT.iterdir()):
        if not subdir.is_dir():
            continue
        for pattern in ("*.xlsx", "*.xls"):
            for f in sorted(subdir.glob(pattern)):
                if not f.name.startswith("~$"):
                    files.append(f)

    if not files:
        print("No spreadsheet files found.")
        return 0

    print(f"Found {len(files)} file(s) to sort:")
    errors = []
    for f in files:
        rel = f.relative_to(ROOT)
        try:
            sort_file(f)
            print(f"  OK  {rel}")
        except Exception as e:
            print(f"  ERR {rel}: {e}")
            errors.append((rel, e))

    if errors:
        print(f"\n{len(errors)} file(s) failed.")
        return 1
    print("\nAll files sorted successfully.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
