#!/usr/bin/env python3
# /// script
# dependencies = ["openpyxl"]
# ///
"""
Concatenate all .txt files in subdirectories of a folder into an Excel file.

Columns: Year | Question Number | Letter | Unit Topic | Source

Usage:
    uv run concat_txt.py <folder>
"""

import argparse
import re
import sys
from pathlib import Path

import openpyxl


def parse_line(line: str) -> tuple | None:
    """Parse a line like '2008B, 1a, Unit 9: ..., Subtopic' into row fields."""
    line = line.strip()
    if not line:
        return None

    # Split into at most 3 parts: year, question+letter, rest
    parts = line.split(", ", maxsplit=2)
    if len(parts) < 3:
        return None

    year = parts[0].strip()
    q_raw = parts[1].strip()
    unit_topic = parts[2].strip()

    # Split question number and letter (e.g. "1a" → "1", "a")
    m = re.fullmatch(r"(\d+)([a-z]*)", q_raw, re.IGNORECASE)
    if not m:
        return None
    number = m.group(1)
    letter = m.group(2)

    return year, number, letter, unit_topic


def main():
    parser = argparse.ArgumentParser(
        description="Concatenate .txt results into an Excel file"
    )
    parser.add_argument("folder", type=Path, help="Root folder to search")
    args = parser.parse_args()

    folder = args.folder.resolve()
    if not folder.is_dir():
        print(f"Error: {folder} is not a directory", file=sys.stderr)
        sys.exit(1)

    txt_files = sorted(folder.rglob("*.txt"))
    if not txt_files:
        print(f"No .txt files found under {folder}", file=sys.stderr)
        sys.exit(1)

    # Prompt user for save location
    default_output = Path.cwd() / "unit-topics.xlsx"
    answer = input(f"Save results to [{default_output}]: ").strip()
    output = Path(answer).expanduser() if answer else default_output
    if output.is_dir():
        output = output / "unit-topics.xlsx"
    if output.suffix.lower() != ".xlsx":
        output = output.with_suffix(".xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Year", "Question Number", "Letter", "Unit Topic", "Source"])

    total_rows = 0
    for txt_file in txt_files:
        pdf_name = txt_file.stem + ".pdf"
        for line in txt_file.read_text(encoding="utf-8").splitlines():
            parsed = parse_line(line)
            if parsed:
                year, number, letter, unit_topic = parsed
                ws.append([year, number, letter, unit_topic, pdf_name])
                cell = ws.cell(row=ws.max_row, column=5)
                cell.hyperlink = pdf_name
                cell.style = "Hyperlink"
                total_rows += 1

    wb.save(output)
    print(f"Wrote {total_rows} rows to {output}")


if __name__ == "__main__":
    main()
